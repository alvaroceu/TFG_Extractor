import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def export_results_to_excel(master_data: list, excel_name="resultados_totales.xlsx"):
    """
    Exports the complete dataset of rows directly to the requested Excel format.
    """
    df = pd.DataFrame(master_data)
    
    # Reordenamos las columnas por seguridad para asegurar el formato pedido
    cols_order = [
        "Question ID", "Dataset", "Context", "Question", "Ground Truth", 
        "Model", "Answer", "ExecTime", "Status", "ExactMatch", 
        "InclusionMatch", "ROUGE_L", "BERTScore"
    ]
    df = df[cols_order]
    
    df.to_excel(excel_name, index=False)
    _format_excel(excel_name)

def _format_excel(excel_name):
    """Función auxiliar con tu código original para poner bonito el Excel."""
    wb = load_workbook(excel_name)
    ws = wb.active

    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 60

    wb.save(excel_name)
